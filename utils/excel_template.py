"""
MALT IAC Excel Template Generator and Reader
Multi-sheet .xlsx with pre-styled input areas, generated via openpyxl.

Sheets:
  1. Cover             — facility/team metadata
  2. Utility_Billing   — electricity, natural gas, water billing tables
  3. Facility_Background — descriptions, best practices, equipment table
  4. AR_Template        — one AR per sheet (supports multiple sheets)
  5. Instructions       — how to fill each sheet
"""

import io
from copy import deepcopy

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ── Color constants ───────────────────────────────────────────────────────────

NAVY_HEX      = "00003366"   # Header rows
MED_BLUE_HEX  = "001565C0"   # Section sub-headers
LIGHT_YELLOW  = "00FFFDE7"   # Editable input cells
LIGHT_GRAY    = "00F5F5F5"   # Alternating data rows
WHITE_HEX     = "00FFFFFF"
LABEL_GRAY    = "00EEEEEE"   # Label cells (non-editable)

NAVY_FILL       = PatternFill("solid", fgColor=NAVY_HEX)
MED_BLUE_FILL   = PatternFill("solid", fgColor=MED_BLUE_HEX)
YELLOW_FILL     = PatternFill("solid", fgColor=LIGHT_YELLOW)
GRAY_FILL       = PatternFill("solid", fgColor=LIGHT_GRAY)
LABEL_FILL      = PatternFill("solid", fgColor=LABEL_GRAY)
WHITE_FILL      = PatternFill("solid", fgColor=WHITE_HEX)

HEADER_FONT     = Font(name="Calibri", bold=True, color="FFFFFFFF", size=11)
SUBHEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFFFF", size=10)
LABEL_FONT      = Font(name="Calibri", bold=True, color="FF333333", size=10)
INPUT_FONT      = Font(name="Calibri", color="FF000000", size=10)
TITLE_FONT      = Font(name="Calibri", bold=True, color="FF003366", size=13)

CENTER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT_ALIGN   = Alignment(horizontal="right",  vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


# ── Style helpers ─────────────────────────────────────────────────────────────

def _header_cell(ws, row: int, col: int, text: str, width_hint: int = None):
    """Write a navy header cell."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill   = NAVY_FILL
    cell.font   = HEADER_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    if width_hint:
        ws.column_dimensions[get_column_letter(col)].width = width_hint
    return cell


def _subheader_cell(ws, row: int, col: int, text: str, span_end_col: int = None):
    """Write a medium-blue sub-header cell, optionally merging columns."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill   = MED_BLUE_FILL
    cell.font   = SUBHEADER_FONT
    cell.alignment = LEFT_ALIGN
    cell.border = THIN_BORDER
    if span_end_col and span_end_col > col:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=span_end_col
        )
    return cell


def _label_cell(ws, row: int, col: int, text: str):
    """Write a gray label cell."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill   = LABEL_FILL
    cell.font   = LABEL_FONT
    cell.alignment = LEFT_ALIGN
    cell.border = THIN_BORDER
    return cell


def _input_cell(ws, row: int, col: int, value=None, named: str = None, num_fmt: str = None):
    """Write a yellow input cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill      = YELLOW_FILL
    cell.font      = INPUT_FONT
    cell.alignment = LEFT_ALIGN
    cell.border    = THIN_BORDER
    if num_fmt:
        cell.number_format = num_fmt
    return cell


def _data_cell(ws, row: int, col: int, value=None, alternate: bool = False, num_fmt: str = None):
    """Write a data cell (alternating row fill)."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill      = GRAY_FILL if alternate else WHITE_FILL
    cell.font      = INPUT_FONT
    cell.alignment = CENTER_ALIGN
    cell.border    = THIN_BORDER
    if num_fmt:
        cell.number_format = num_fmt
    return cell


def _set_col_widths(ws, widths: dict):
    """widths: {col_letter_or_int: width_float}"""
    for col, w in widths.items():
        if isinstance(col, int):
            col = get_column_letter(col)
        ws.column_dimensions[col].width = w


def _row_height(ws, row: int, height: float):
    ws.row_dimensions[row].height = height


# ── Sheet 1: Cover ────────────────────────────────────────────────────────────

def _build_cover_sheet(ws, session: dict = None):
    s = session or {}

    ws.title = "Cover"
    ws.sheet_view.showGridLines = True

    # Title
    title_cell = ws.cell(row=1, column=1, value="MALT IAC Energy Assessment — Cover Information")
    title_cell.font = TITLE_FONT
    title_cell.alignment = LEFT_ALIGN
    ws.merge_cells("A1:D1")
    _row_height(ws, 1, 22)

    fields = [
        ("Report_Number",      "Report Number",            s.get("report_number", "")),
        ("Site_Visit_Date",    "Site Visit Date",          str(s.get("site_visit_date", "")) if s.get("site_visit_date") else ""),
        ("Facility_Name",      "Facility Name",            s.get("facility_name", s.get("location", "").split(",")[0].strip() if s.get("location") else "")),
        ("City",               "City",                     ""),
        ("State",              "State",                    ""),
        ("Zip",                "ZIP Code",                 ""),
        ("NAICS_Code",         "NAICS Code",               s.get("naics_code", "")),
        ("SIC_Code",           "SIC Code",                 s.get("sic_code", "")),
        ("Principal_Products", "Principal Products",       s.get("principal_products", "")),
        ("Floor_Area_ft2",     "Floor Area (sq ft)",       ""),
        ("Director_Name",      "IAC Director / Lead Faculty", s.get("lead_faculty", "")),
        ("Lead_Faculty",       "Lead Faculty",             s.get("lead_faculty", "")),
        ("Lead_Student",       "Lead Student",             s.get("lead_student", "")),
        ("Safety_Student",     "Safety Student",           s.get("safety_student", "")),
        ("Assessment_Team",    "Assessment Team (comma-separated)", s.get("other_students", "")),
        ("Annual_Sales",       "Annual Sales ($)",         s.get("annual_sales", "")),
        ("Num_Employees",      "Number of Employees",      s.get("num_employees", "")),
    ]

    for i, (key, label, value) in enumerate(fields):
        row = i + 3
        _label_cell(ws, row, 1, label)
        _input_cell(ws, row, 2, value)
        # Named range hint in col C
        ws.cell(row=row, column=3, value=f"← {key}").font = Font(name="Calibri", size=9, color="FF888888", italic=True)

    _set_col_widths(ws, {1: 28, 2: 32, 3: 20, 4: 14})

    # Instructions
    inst_row = len(fields) + 4
    ws.cell(row=inst_row, column=1,
            value="Fill all yellow cells. Date format: YYYY-MM-DD").font = Font(
        name="Calibri", size=9, italic=True, color="FF666666")


# ── Sheet 2: Utility_Billing ──────────────────────────────────────────────────

def _build_utility_sheet(ws, session: dict = None):
    s = session or {}
    ws.title = "Utility_Billing"

    elec_rows  = s.get("elec_rows",  [{"month": m, "kwh": 0, "elec_cost": 0, "kw": 0, "demand_cost": 0, "fee": 0, "total": 0} for m in MONTHS])
    gas_rows   = s.get("gas_rows",   [{"month": m, "mmbtu": 0, "cost": 0, "fee": 0, "total": 0} for m in MONTHS])
    water_rows = s.get("water_rows", [{"month": m, "tgal": 0, "water_cost": 0, "sewer_cost": 0, "fee": 0, "total": 0} for m in MONTHS])

    cur_row = 1

    # ── Electricity ──
    _subheader_cell(ws, cur_row, 1, "ELECTRICITY BILLING", span_end_col=7)
    cur_row += 1

    elec_headers = ["Month", "kWh", "Electricity Cost ($)", "Demand kW", "Demand Cost ($)", "Fees ($)", "Total ($)"]
    for c, h in enumerate(elec_headers, 1):
        _header_cell(ws, cur_row, c, h)
    cur_row += 1
    elec_data_start = cur_row

    for i, r in enumerate(elec_rows):
        alt = i % 2 == 1
        month_val = r.get("month", MONTHS[i] if i < 12 else "")
        _label_cell(ws, cur_row, 1, month_val)
        ws.cell(row=cur_row, column=1).fill = GRAY_FILL if alt else WHITE_FILL
        _input_cell(ws, cur_row, 2, r.get("kwh", 0) or None, num_fmt="#,##0.00")
        _input_cell(ws, cur_row, 3, r.get("elec_cost", 0) or None, num_fmt="#,##0.00")
        _input_cell(ws, cur_row, 4, r.get("kw", 0) or None, num_fmt="#,##0.00")
        _input_cell(ws, cur_row, 5, r.get("demand_cost", 0) or None, num_fmt="#,##0.00")
        _input_cell(ws, cur_row, 6, r.get("fee", 0) or None, num_fmt="#,##0.00")
        # Total formula
        col_b, col_c, col_e, col_f = "B", "C", "E", "F"
        total_cell = ws.cell(row=cur_row, column=7,
                              value=f"=IFERROR({col_c}{cur_row}+{col_e}{cur_row}+{col_f}{cur_row},0)")
        total_cell.fill   = GRAY_FILL if alt else WHITE_FILL
        total_cell.font   = INPUT_FONT
        total_cell.border = THIN_BORDER
        total_cell.number_format = "#,##0.00"
        total_cell.alignment = CENTER_ALIGN
        cur_row += 1

    # Totals row for electricity
    _label_cell(ws, cur_row, 1, "TOTAL")
    for c in range(2, 8):
        col_letter = get_column_letter(c)
        tc = ws.cell(row=cur_row, column=c,
                     value=f"=SUM({col_letter}{elec_data_start}:{col_letter}{cur_row-1})")
        tc.fill   = MED_BLUE_FILL
        tc.font   = Font(name="Calibri", bold=True, color="FFFFFFFF", size=10)
        tc.border = THIN_BORDER
        tc.number_format = "#,##0.00"
        tc.alignment = CENTER_ALIGN
    cur_row += 2

    # ── Natural Gas ──
    _subheader_cell(ws, cur_row, 1, "NATURAL GAS BILLING", span_end_col=5)
    cur_row += 1

    gas_headers = ["Month", "MMBtu", "Gas Cost ($)", "Fees ($)", "Total ($)"]
    for c, h in enumerate(gas_headers, 1):
        _header_cell(ws, cur_row, c, h)
    cur_row += 1
    gas_data_start = cur_row

    for i, r in enumerate(gas_rows):
        alt = i % 2 == 1
        month_val = r.get("month", MONTHS[i] if i < 12 else "")
        _label_cell(ws, cur_row, 1, month_val)
        ws.cell(row=cur_row, column=1).fill = GRAY_FILL if alt else WHITE_FILL
        _input_cell(ws, cur_row, 2, r.get("mmbtu", 0) or None, num_fmt="#,##0.000")
        _input_cell(ws, cur_row, 3, r.get("cost", 0) or None, num_fmt="#,##0.00")
        _input_cell(ws, cur_row, 4, r.get("fee", 0) or None, num_fmt="#,##0.00")
        total_cell = ws.cell(row=cur_row, column=5,
                              value=f"=IFERROR(C{cur_row}+D{cur_row},0)")
        total_cell.fill   = GRAY_FILL if alt else WHITE_FILL
        total_cell.font   = INPUT_FONT
        total_cell.border = THIN_BORDER
        total_cell.number_format = "#,##0.00"
        total_cell.alignment = CENTER_ALIGN
        cur_row += 1

    _label_cell(ws, cur_row, 1, "TOTAL")
    for c in range(2, 6):
        col_letter = get_column_letter(c)
        tc = ws.cell(row=cur_row, column=c,
                     value=f"=SUM({col_letter}{gas_data_start}:{col_letter}{cur_row-1})")
        tc.fill   = MED_BLUE_FILL
        tc.font   = Font(name="Calibri", bold=True, color="FFFFFFFF", size=10)
        tc.border = THIN_BORDER
        tc.number_format = "#,##0.00"
        tc.alignment = CENTER_ALIGN
    cur_row += 2

    # ── Water ──
    _subheader_cell(ws, cur_row, 1, "WATER / SEWER BILLING", span_end_col=6)
    cur_row += 1

    water_headers = ["Month", "Tgal", "Water Cost ($)", "Sewer Cost ($)", "Fees ($)", "Total ($)"]
    for c, h in enumerate(water_headers, 1):
        _header_cell(ws, cur_row, c, h)
    cur_row += 1
    water_data_start = cur_row

    for i, r in enumerate(water_rows):
        alt = i % 2 == 1
        month_val = r.get("month", MONTHS[i] if i < 12 else "")
        _label_cell(ws, cur_row, 1, month_val)
        ws.cell(row=cur_row, column=1).fill = GRAY_FILL if alt else WHITE_FILL
        _input_cell(ws, cur_row, 2, r.get("tgal", 0) or None, num_fmt="#,##0.0000")
        _input_cell(ws, cur_row, 3, r.get("water_cost", 0) or None, num_fmt="#,##0.00")
        _input_cell(ws, cur_row, 4, r.get("sewer_cost", 0) or None, num_fmt="#,##0.00")
        _input_cell(ws, cur_row, 5, r.get("fee", 0) or None, num_fmt="#,##0.00")
        total_cell = ws.cell(row=cur_row, column=6,
                              value=f"=IFERROR(C{cur_row}+D{cur_row}+E{cur_row},0)")
        total_cell.fill   = GRAY_FILL if alt else WHITE_FILL
        total_cell.font   = INPUT_FONT
        total_cell.border = THIN_BORDER
        total_cell.number_format = "#,##0.00"
        total_cell.alignment = CENTER_ALIGN
        cur_row += 1

    _label_cell(ws, cur_row, 1, "TOTAL")
    for c in range(2, 7):
        col_letter = get_column_letter(c)
        tc = ws.cell(row=cur_row, column=c,
                     value=f"=SUM({col_letter}{water_data_start}:{col_letter}{cur_row-1})")
        tc.fill   = MED_BLUE_FILL
        tc.font   = Font(name="Calibri", bold=True, color="FFFFFFFF", size=10)
        tc.border = THIN_BORDER
        tc.number_format = "#,##0.00"
        tc.alignment = CENTER_ALIGN

    _set_col_widths(ws, {1: 10, 2: 14, 3: 18, 4: 16, 5: 14, 6: 14, 7: 14})


# ── Sheet 3: Facility_Background ─────────────────────────────────────────────

def _build_facility_sheet(ws, session: dict = None):
    s = session or {}
    ws.title = "Facility_Background"

    cur_row = 1

    # Title
    title = ws.cell(row=cur_row, column=1, value="MALT IAC — Facility Background")
    title.font = TITLE_FONT
    title.alignment = LEFT_ALIGN
    ws.merge_cells(f"A{cur_row}:E{cur_row}")
    _row_height(ws, cur_row, 20)
    cur_row += 2

    # Facility Description
    _subheader_cell(ws, cur_row, 1, "Facility Description", span_end_col=5)
    cur_row += 1
    cell = _input_cell(ws, cur_row, 1, s.get("facility_description", ""))
    ws.merge_cells(f"A{cur_row}:E{cur_row}")
    ws.row_dimensions[cur_row].height = 60
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cur_row += 2

    # Process Description
    _subheader_cell(ws, cur_row, 1, "Process Description", span_end_col=5)
    cur_row += 1
    cell = _input_cell(ws, cur_row, 1, s.get("process_description", ""))
    ws.merge_cells(f"A{cur_row}:E{cur_row}")
    ws.row_dimensions[cur_row].height = 60
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cur_row += 2

    # Best Practices (3 cells)
    _subheader_cell(ws, cur_row, 1, "Best Practices (enter one per cell)", span_end_col=5)
    cur_row += 1
    best_practices = s.get("best_practices", ["", "", ""])
    for bp_idx in range(3):
        _label_cell(ws, cur_row, 1, f"Best Practice {bp_idx + 1}:")
        bp_val = best_practices[bp_idx] if bp_idx < len(best_practices) else ""
        cell = _input_cell(ws, cur_row, 2, bp_val)
        ws.merge_cells(f"B{cur_row}:E{cur_row}")
        cur_row += 1
    cur_row += 1

    # Energy Forms
    _subheader_cell(ws, cur_row, 1, "Forms of Energy Usage", span_end_col=5)
    cur_row += 1
    _label_cell(ws, cur_row, 1, "Electricity Used For:")
    elec_uses = s.get("elec_used_for", [])
    cell = _input_cell(ws, cur_row, 2, ", ".join(elec_uses) if elec_uses else "")
    ws.merge_cells(f"B{cur_row}:E{cur_row}")
    cur_row += 1
    _label_cell(ws, cur_row, 1, "Natural Gas Used For:")
    gas_uses = s.get("gas_used_for", [])
    cell = _input_cell(ws, cur_row, 2, ", ".join(gas_uses) if gas_uses else "")
    ws.merge_cells(f"B{cur_row}:E{cur_row}")
    cur_row += 2

    # Equipment Table
    _subheader_cell(ws, cur_row, 1, "Major Energy Consuming Equipment", span_end_col=5)
    cur_row += 1
    eq_headers = ["#", "Equipment / System", "Specifications", "Qty / Capacity", "Energy Form"]
    for c, h in enumerate(eq_headers, 1):
        _header_cell(ws, cur_row, c, h)
    cur_row += 1

    eq_rows = s.get("equipment_rows", [])
    valid_eq = [r for r in eq_rows if r.get("equipment", "")]
    if not valid_eq:
        # Provide 10 blank rows
        for i in range(10):
            alt = i % 2 == 1
            _data_cell(ws, cur_row, 1, i + 1, alt)
            for c in range(2, 6):
                _input_cell(ws, cur_row, c)
            cur_row += 1
    else:
        for i, r in enumerate(valid_eq):
            alt = i % 2 == 1
            _data_cell(ws, cur_row, 1, i + 1, alt)
            _input_cell(ws, cur_row, 2, r.get("equipment", ""))
            _input_cell(ws, cur_row, 3, r.get("specs", ""))
            _input_cell(ws, cur_row, 4, r.get("qty_capacity", ""))
            _input_cell(ws, cur_row, 5, r.get("energy_form", ""))
            cur_row += 1
        # Add a few blank rows for additions
        filled = len(valid_eq)
        for i in range(5):
            alt = (filled + i) % 2 == 1
            _data_cell(ws, cur_row, 1, filled + i + 1, alt)
            for c in range(2, 6):
                _input_cell(ws, cur_row, c)
            cur_row += 1

    _set_col_widths(ws, {1: 5, 2: 28, 3: 28, 4: 18, 5: 16})


# ── Sheet 4: AR_Template ──────────────────────────────────────────────────────

def _build_ar_sheet(ws, ar_data: dict = None, ar_idx: int = 1):
    """Build one AR sheet. ar_data is an AR dict from session ar_list."""
    ar = ar_data or {}
    ws.title = f"AR_{ar_idx}"

    cur_row = 1

    # Title
    title_val = f"Assessment Recommendation #{ar_idx}"
    title = ws.cell(row=cur_row, column=1, value=title_val)
    title.font = TITLE_FONT
    title.alignment = LEFT_ALIGN
    ws.merge_cells(f"A{cur_row}:D{cur_row}")
    _row_height(ws, cur_row, 22)
    cur_row += 2

    # Identity fields
    identity_fields = [
        ("AR_Number",   "AR Number",    ar.get("ar_number", f"AR-{ar_idx}")),
        ("ARC_Code",    "ARC Code",     ar.get("arc_code", "")),
        ("Title",       "Title",        ar.get("title", "")),
    ]
    for key, label, value in identity_fields:
        _label_cell(ws, cur_row, 1, label)
        cell = _input_cell(ws, cur_row, 2, value)
        ws.merge_cells(f"B{cur_row}:D{cur_row}")
        cur_row += 1
    cur_row += 1

    # Text narrative fields
    narrative_fields = [
        ("Observation",            "Observation",             ar.get("observation", "")),
        ("Recommendation",         "Recommendation",          ar.get("recommendation", "")),
        ("Technology_Description", "Technology Description",  ar.get("tech_description", "")),
    ]
    for key, label, value in narrative_fields:
        _subheader_cell(ws, cur_row, 1, label, span_end_col=4)
        cur_row += 1
        cell = _input_cell(ws, cur_row, 1, value)
        ws.merge_cells(f"A{cur_row}:D{cur_row}")
        ws.row_dimensions[cur_row].height = 72
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cur_row += 2

    # Resource savings
    _subheader_cell(ws, cur_row, 1, "Resource Savings", span_end_col=4)
    cur_row += 1
    res_headers = ["Resource Type", "Annual Savings", "Unit"]
    for c, h in enumerate(res_headers, 1):
        _header_cell(ws, cur_row, c, h)
    cur_row += 1

    resources = ar.get("resources", [])
    for r_idx in range(max(2, len(resources))):
        r = resources[r_idx] if r_idx < len(resources) else {}
        _input_cell(ws, cur_row, 1, r.get("type", ""))
        _input_cell(ws, cur_row, 2, r.get("savings", None) if r.get("savings", 0) != 0 else None,
                    num_fmt="#,##0.00")
        _input_cell(ws, cur_row, 3, r.get("unit", ""))
        cur_row += 1
    cur_row += 1

    # Financial summary
    _subheader_cell(ws, cur_row, 1, "Financial Summary", span_end_col=4)
    cur_row += 1
    fin_fields = [
        ("Total_Cost_Savings",   "Total Annual Cost Savings ($)",  ar.get("total_cost_savings", None)),
        ("Implementation_Cost",  "Implementation Cost ($)",        ar.get("implementation_cost", None)),
        ("Payback_Years",        "Simple Payback Period (years)",  ar.get("payback", None)),
    ]
    for key, label, value in fin_fields:
        _label_cell(ws, cur_row, 1, label)
        val = value if (value is not None and value != float("inf")) else None
        _input_cell(ws, cur_row, 2, val, num_fmt="#,##0.00")
        cur_row += 1
    cur_row += 1

    # Calculation details
    _subheader_cell(ws, cur_row, 1, "Calculation Details (key: value pairs)", span_end_col=4)
    cur_row += 1
    _header_cell(ws, cur_row, 1, "Parameter")
    _header_cell(ws, cur_row, 2, "Value")
    cur_row += 1

    calc = ar.get("calculation_details", {})
    if calc:
        for k, v in calc.items():
            if k == "narrative":
                continue
            _input_cell(ws, cur_row, 1, k.replace("_", " ").title())
            val_str = f"{v:,.4f}" if isinstance(v, float) else str(v)
            _input_cell(ws, cur_row, 2, val_str)
            cur_row += 1
    else:
        for _ in range(8):
            _input_cell(ws, cur_row, 1)
            _input_cell(ws, cur_row, 2)
            cur_row += 1

    _set_col_widths(ws, {1: 30, 2: 22, 3: 14, 4: 14})


# ── Sheet 5: Instructions ─────────────────────────────────────────────────────

def _build_instructions_sheet(ws):
    ws.title = "Instructions"

    instructions = [
        ("MALT IAC Excel Template — Instructions", True, 14, "00003366"),
        ("", False, 11, "00000000"),
        ("HOW TO USE THIS TEMPLATE", True, 12, "001565C0"),
        ("1. Fill all yellow-highlighted cells with your data.", False, 11, "00000000"),
        ("2. Gray cells with labels are not editable — they identify field names.", False, 11, "00000000"),
        ("3. Blue header cells identify sections — do not edit.", False, 11, "00000000"),
        ("4. After filling all sheets, save as .xlsx and upload to the MALT tool.", False, 11, "00000000"),
        ("", False, 11, "00000000"),
        ("SHEET GUIDE", True, 12, "001565C0"),
        ("Cover:              Fill in facility identification, location, team, and NAICS/SIC codes.", False, 11, "00000000"),
        ("Utility_Billing:    Enter 12 months of electricity, natural gas, and water billing data.", False, 11, "00000000"),
        ("                    Totals are calculated automatically via Excel formulas.", False, 11, "00444444"),
        ("Facility_Background: Enter facility/process descriptions, best practices, energy uses,", False, 11, "00000000"),
        ("                    and major equipment list.", False, 11, "00444444"),
        ("AR_1, AR_2, ...:    Each sheet holds one Assessment Recommendation.", False, 11, "00000000"),
        ("                    Fill observation, recommendation, technology description,", False, 11, "00444444"),
        ("                    resource savings, and financial summary.", False, 11, "00444444"),
        ("", False, 11, "00000000"),
        ("FIELD NAMING CONVENTION", True, 12, "001565C0"),
        ("The gray 'key' column on the Cover sheet shows the internal field names", False, 11, "00000000"),
        ("used by the tool. Do not change them — they link Excel data to the app.", False, 11, "00000000"),
        ("", False, 11, "00000000"),
        ("DATE FORMAT", True, 12, "001565C0"),
        ("Enter dates as YYYY-MM-DD (e.g., 2025-10-15) for reliable import.", False, 11, "00000000"),
        ("", False, 11, "00000000"),
        ("QUESTIONS", True, 12, "001565C0"),
        ("Contact the MALT IAC at the University of Louisiana at Lafayette.", False, 11, "00000000"),
    ]

    for row_idx, (text, bold, size, color) in enumerate(instructions, 1):
        cell = ws.cell(row=row_idx, column=1, value=text)
        cell.font = Font(name="Calibri", bold=bold, size=size, color=f"FF{color[2:]}")
        cell.alignment = LEFT_ALIGN
        _row_height(ws, row_idx, 16 if size == 11 else 22)

    ws.column_dimensions["A"].width = 80


# ── Main: generate_excel_template ─────────────────────────────────────────────

def generate_excel_template(session: dict = None) -> bytes:
    """
    Generate blank (or pre-filled) Excel template.
    Returns bytes suitable for st.download_button.
    """
    s = session or {}

    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Build standard sheets
    _build_cover_sheet(wb.create_sheet("Cover"), s)
    _build_utility_sheet(wb.create_sheet("Utility_Billing"), s)
    _build_facility_sheet(wb.create_sheet("Facility_Background"), s)

    # AR sheets — one per AR in session, or at least one blank template
    ar_list = s.get("ar_list", [])
    if ar_list:
        for idx, ar in enumerate(ar_list):
            sheet_name = f"AR_{idx + 1}"
            _build_ar_sheet(wb.create_sheet(sheet_name), ar, idx + 1)
    else:
        _build_ar_sheet(wb.create_sheet("AR_Template"), {}, 1)

    _build_instructions_sheet(wb.create_sheet("Instructions"))

    # Set Cover as active sheet
    wb.active = wb["Cover"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── read_excel_template ───────────────────────────────────────────────────────

def read_excel_template(file_bytes: bytes) -> dict:
    """
    Read a filled MALT Excel template and return a session-compatible dict.
    Maps cell values back to session state keys.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = {}

    # ── Cover sheet ──────────────────────────────────────────────────────────
    if "Cover" in wb.sheetnames:
        ws = wb["Cover"]
        # Key mapping: row offset from row 3, column B (col 2)
        cover_keys = [
            "report_number", "site_visit_date", "facility_name",
            "city", "state", "zip_code",
            "naics_code", "sic_code", "principal_products",
            "floor_area_ft2",
            "director_name", "lead_faculty",
            "lead_student", "safety_student", "other_students",
            "annual_sales", "num_employees",
        ]
        # Rows start at 3
        for i, key in enumerate(cover_keys):
            row = i + 3
            val = ws.cell(row=row, column=2).value
            if val is not None and str(val).strip():
                result[key] = str(val).strip()
        # Reconstruct location from parsed cover fields
        facility_name = result.get("facility_name", "")
        city = result.get("city", "")
        state = result.get("state", "")
        zip_code = result.get("zip_code", "")
        if facility_name or city:
            parts = [p for p in [facility_name, city, state, zip_code] if p]
            result["location"] = ", ".join(parts)

    # ── Utility_Billing sheet ─────────────────────────────────────────────────
    if "Utility_Billing" in wb.sheetnames:
        ws = wb["Utility_Billing"]

        def _safe_float(v):
            try:
                return float(v) if v is not None else 0.0
            except (ValueError, TypeError):
                return 0.0

        # Find section start rows by scanning for subheader text
        elec_start = gas_start = water_start = None
        for row in ws.iter_rows():
            for cell in row:
                val = str(cell.value or "").upper()
                if "ELECTRICITY" in val and cell.column == 1:
                    elec_start = cell.row
                elif "NATURAL GAS" in val and cell.column == 1:
                    gas_start = cell.row
                elif "WATER" in val and cell.column == 1:
                    water_start = cell.row

        # Read electricity (header row +1, then 12 data rows)
        if elec_start:
            data_row = elec_start + 2  # subheader + header = +2
            elec_rows = []
            for m_idx, month in enumerate(MONTHS):
                r_vals = [ws.cell(row=data_row + m_idx, column=c).value for c in range(1, 8)]
                elec_rows.append({
                    "month":       str(r_vals[0] or month),
                    "kwh":         _safe_float(r_vals[1]),
                    "elec_cost":   _safe_float(r_vals[2]),
                    "kw":          _safe_float(r_vals[3]),
                    "demand_cost": _safe_float(r_vals[4]),
                    "fee":         _safe_float(r_vals[5]),
                    "total":       _safe_float(r_vals[6]),
                })
            result["elec_rows"] = elec_rows

        # Read natural gas
        if gas_start:
            data_row = gas_start + 2
            gas_rows = []
            for m_idx, month in enumerate(MONTHS):
                r_vals = [ws.cell(row=data_row + m_idx, column=c).value for c in range(1, 6)]
                gas_rows.append({
                    "month":  str(r_vals[0] or month),
                    "mmbtu":  _safe_float(r_vals[1]),
                    "cost":   _safe_float(r_vals[2]),
                    "fee":    _safe_float(r_vals[3]),
                    "total":  _safe_float(r_vals[4]),
                })
            result["gas_rows"] = gas_rows

        # Read water
        if water_start:
            data_row = water_start + 2
            water_rows = []
            for m_idx, month in enumerate(MONTHS):
                r_vals = [ws.cell(row=data_row + m_idx, column=c).value for c in range(1, 7)]
                water_rows.append({
                    "month":      str(r_vals[0] or month),
                    "tgal":       _safe_float(r_vals[1]),
                    "water_cost": _safe_float(r_vals[2]),
                    "sewer_cost": _safe_float(r_vals[3]),
                    "fee":        _safe_float(r_vals[4]),
                    "total":      _safe_float(r_vals[5]),
                })
            result["water_rows"] = water_rows

    # ── Facility_Background sheet ─────────────────────────────────────────────
    if "Facility_Background" in wb.sheetnames:
        ws = wb["Facility_Background"]
        all_rows = list(ws.iter_rows(values_only=True))

        def _find_section_row(keyword: str) -> int:
            for i, row in enumerate(all_rows):
                if row[0] and keyword.upper() in str(row[0]).upper():
                    return i
            return -1

        # Facility description (row after its header)
        fd_row = _find_section_row("Facility Description")
        if fd_row >= 0 and fd_row + 1 < len(all_rows):
            val = all_rows[fd_row + 1][0]
            if val:
                result["facility_description"] = str(val)

        pd_row = _find_section_row("Process Description")
        if pd_row >= 0 and pd_row + 1 < len(all_rows):
            val = all_rows[pd_row + 1][0]
            if val:
                result["process_description"] = str(val)

        bp_row = _find_section_row("Best Practices")
        if bp_row >= 0:
            bps = []
            for offset in range(1, 4):
                r = all_rows[bp_row + offset] if bp_row + offset < len(all_rows) else ()
                # Best practice text is in column B (index 1)
                val = r[1] if len(r) > 1 else None
                bps.append(str(val) if val else "")
            result["best_practices"] = bps

        ef_row = _find_section_row("Forms of Energy")
        if ef_row >= 0:
            # Electricity uses
            e_row = all_rows[ef_row + 1] if ef_row + 1 < len(all_rows) else ()
            e_val = e_row[1] if len(e_row) > 1 else None
            if e_val:
                result["elec_used_for"] = [v.strip() for v in str(e_val).split(",") if v.strip()]
            # Gas uses
            g_row = all_rows[ef_row + 2] if ef_row + 2 < len(all_rows) else ()
            g_val = g_row[1] if len(g_row) > 1 else None
            if g_val:
                result["gas_used_for"] = [v.strip() for v in str(g_val).split(",") if v.strip()]

        eq_row = _find_section_row("Major Energy")
        if eq_row >= 0:
            # Skip section header + column header row
            eq_data_start = eq_row + 2
            equipment_rows = []
            for offset in range(0, 30):
                r_idx = eq_data_start + offset
                if r_idx >= len(all_rows):
                    break
                r = all_rows[r_idx]
                if not r or not r[1]:  # col B = equipment name
                    break
                equipment_rows.append({
                    "equipment":   str(r[1] or ""),
                    "specs":       str(r[2] or ""),
                    "qty_capacity": str(r[3] or ""),
                    "energy_form": str(r[4] or "Electricity"),
                })
            if equipment_rows:
                result["equipment_rows"] = equipment_rows

    # ── AR sheets ─────────────────────────────────────────────────────────────
    ar_sheets = [name for name in wb.sheetnames
                 if name.startswith("AR_") and name != "AR_Template"]
    if ar_sheets:
        ar_list_out = []
        for sheet_name in sorted(ar_sheets):
            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))

            def _get_labeled_val(label_keyword: str):
                for row in all_rows:
                    if row[0] and label_keyword.lower() in str(row[0]).lower():
                        return row[1] if len(row) > 1 else None
                return None

            def _get_text_after_header(keyword: str):
                for i, row in enumerate(all_rows):
                    if row[0] and keyword.lower() in str(row[0]).lower():
                        if i + 1 < len(all_rows):
                            return all_rows[i + 1][0]
                return None

            def _safe_float_ar(v):
                try:
                    return float(v) if v is not None else 0.0
                except (ValueError, TypeError):
                    return 0.0

            ar_number  = _get_labeled_val("AR Number") or ""
            arc_code   = _get_labeled_val("ARC Code") or ""
            title_val  = _get_labeled_val("Title") or ""
            observation  = _get_text_after_header("Observation") or ""
            recommendation = _get_text_after_header("Recommendation") or ""
            tech_desc    = _get_text_after_header("Technology Description") or ""
            cost_sav   = _safe_float_ar(_get_labeled_val("Annual Cost Savings"))
            impl_cost  = _safe_float_ar(_get_labeled_val("Implementation Cost"))
            payback    = _safe_float_ar(_get_labeled_val("Payback"))

            # Resources: scan for resource section
            resources = []
            res_header_row = -1
            for i, row in enumerate(all_rows):
                if row[0] and "Resource Type" in str(row[0]):
                    res_header_row = i
                    break
            if res_header_row >= 0:
                for offset in range(1, 10):
                    r = all_rows[res_header_row + offset] if res_header_row + offset < len(all_rows) else ()
                    if not r or not r[0]:
                        break
                    label_str = str(r[0]).strip()
                    # Stop if we've hit the Financial Summary section
                    if "financial" in label_str.lower() or "total" in label_str.lower():
                        break
                    resources.append({
                        "type":    label_str,
                        "savings": _safe_float_ar(r[1] if len(r) > 1 else None),
                        "unit":    str(r[2]) if len(r) > 2 and r[2] else "",
                    })

            ar_dict = {
                "ar_number":           str(ar_number),
                "arc_code":            str(arc_code),
                "title":               str(title_val),
                "observation":         str(observation),
                "recommendation":      str(recommendation),
                "tech_description":    str(tech_desc),
                "resources":           resources,
                "total_cost_savings":  cost_sav,
                "implementation_cost": impl_cost,
                "payback":             payback if payback > 0 else float("inf"),
                "calculation_details": {},
            }
            # Only add if there's at least an AR number or title
            if ar_number or title_val:
                ar_list_out.append(ar_dict)

        if ar_list_out:
            result["ar_list"] = ar_list_out

    return result
